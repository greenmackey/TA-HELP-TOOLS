
import gspread
from oauth2client.service_account import ServiceAccountCredentials

import openpyxl

# スプレッドシートの鍵を取得
# （例）"10tTyg...5p8k"などの文字列
# jsonとかよくわからんかったら，spread_sheet_keyに値をそのままいれてもいいです．
import json
json_open = open('gspread_info.json', 'r')
json_load = json.load(json_open)
spread_sheet_key = json_load["GSPREAD_SHEET_KEY"]

# jsonfile名を取得
# （例）"coremicro-ii-grade-88c...046.json"などの文字列
# jsonとかよくわからなかったら，jsonfに値をそのままいれてもいいです．
jsonf = json_load["GSPREAD_JSON_FILE_NAME"]

# スプレッドシートでスコアを入力する先のシート番号（1始まり）
print("スプレッドシートでスコアを入力する先のシート番号（1始まり）を入力してください: ", end="")
sheet_num = int(input()) - 1

# スプレッドシートでスコアを入力する先の列番号（アルファベット）
print("スプレッドシートでスコアを入力する先の列番号（アルファベット）を入力してください: ", end="")
col_num_to = openpyxl.utils.column_index_from_string(input())

# 取得先のスプレッドシートで上から数えて無視する行数
print("スプレッドシートで上から数えて無視する行数を入力してください: ", end="")
gspread_row_offset = int(input())

# スコアを取得するのエクセルファイルの絶対パス
# 検索すれば絶対パスを簡単に取得する方法が見つかります
print("スコアを取得するエクセルファイルの絶対パスを入力してください（e.g., /Users/hoge/Desktop/coremicro/hogehoge.xlsx) :")
excel_file_path = input()

# エクセルファイルでスコアが入力されている列番号（アルファベット）
print("エクセルファイルでスコアが入力されている列番号（アルファベット）を入力してください: ", end="")
col_index_from = input()

# エクセルファイルで上から数えて無視する行数
print("書き込み先のエクセルファイルで上から数えて無視する行数を入力してください: ", end="")
excel_row_offset = int(input())




# 取得するエクセルファイルのworkbook, worksheetにアクセス
# 計算結果がほしいのでdata_only=True
excel_wb = openpyxl.load_workbook(excel_file_path, data_only=True)
excel_ws = excel_wb[excel_wb.sheetnames[0]]

# エクセルファイルから学生証番号を取得
excel_student_id_list = excel_ws['A']

# エクセルファイルからスコアを取得
excel_score_cell_list = excel_ws[col_index_from]




#Google Spread Sheetsにアクセス
def connect_gspread(jsonf,key):
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name(jsonf, scope)
    gc = gspread.authorize(credentials)
    SPREADSHEET_KEY = key
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    return workbook

# workbookを取得
wb = connect_gspread(jsonf,spread_sheet_key)

# 指定されたworksheetを取得
ws = wb.get_worksheet(sheet_num)

# スプレッドシートの行数
gspread_row_count = ws.row_count

# 学生証番号を取得
# ヘッダーがあることに注意
gspread_student_id_list = ws.col_values(1)


# スコアを書き込むセルを取得
# ヘッダーが含まれることに注意
gspread_score_cell_list = ws.range(1, col_num_to, gspread_row_count, col_num_to)



for excel_cell_id in range(excel_row_offset, len(excel_score_cell_list)):
  excel_cell = excel_score_cell_list[excel_cell_id]
  excel_row = excel_cell.row
  excel_student_id = excel_ws.cell(row=excel_row, column=1).value
  for gspread_id in range(gspread_row_offset, len(gspread_student_id_list)):
    gspread_student_id = gspread_student_id_list[gspread_id]
    if excel_student_id == gspread_student_id:
      gspread_cell = gspread_score_cell_list[gspread_id]
      gspread_cell.value = excel_cell.value

ws.update_cells(gspread_score_cell_list)


