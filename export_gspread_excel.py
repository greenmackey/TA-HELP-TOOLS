
import gspread
from oauth2client.service_account import ServiceAccountCredentials

import openpyxl

# スプレッドシートの鍵を取得
# （例）"10tTyg...5p8k"などの文字列
# jsonとかよくわからんかったら，spread_sheet_keyに値をそのままいれてもいいです．
import json
json_open = open('gspread_info.json', 'r')
json_load = json.load(json_open)
spread_sheet_key = json_load["GSPREAD_SHEET_KEY"]

# jsonfile名を取得
# （例）"coremicro-ii-grade-88c...046.json"などの文字列
# jsonとかよくわからなかったら，jsonfに値をそのままいれてもいいです．
jsonf = json_load["GSPREAD_JSON_FILE_NAME"]


# エクセルの学生証番号が入っている列番号（アルファベット）
excel_student_id_col_index = "E"
# スプレッドシートの学生証番号が入っている列
gspread_student_id_col_num = 1



# スプレッドシートで参照するシート番号（1始まり）
print("スプレッドシートで参照するのシート番号（1始まり）を入力してください: ", end="")
sheet_num = int(input()) - 1

# スプレッドシートで参照するの最初列番号（アルファベット）
print("スプレッドシートで参照する最初の列番号（アルファベット）を入力してください: ", end="")
col_num_to_start = openpyxl.utils.column_index_from_string(input())

# スプレッドシートでスコアを入力する先の最初列番号（アルファベット）
print("スプレッドシートで参照するの最後の列番号（アルファベット）を入力してください: ", end="")
col_num_to_end = openpyxl.utils.column_index_from_string(input())

# スプレッドシートで上から数えて無視する行数
print("スプレッドシートで上から数えて無視する行数を入力してください: ", end="")
gspread_row_offset = int(input())

# エクセルファイルの絶対パス
# 検索すれば絶対パスを簡単に取得する方法が見つかります
print("エクセルファイルの絶対パスを入力してください（e.g., /Users/hoge/Desktop/coremicro/hogehoge.xlsx) :")
excel_file_path = input()

# エクセルファイルでスコアが入力されている最初の列番号（アルファベット）
print("エクセルファイルで参照する最初の列番号（アルファベット）を入力してください: ", end="")
col_index_from_start = input()
# col_num_from_start = openpyxl.utils.column_index_from_string(col_index_from_start)

# エクセルファイルでスコアが入力されている最後の列番号（アルファベット）
print("エクセルファイルで参照する最後の列番号（アルファベット）を入力してください: ", end="")
col_index_from_end = input()
# col_num_from_end = openpyxl.utils.column_index_from_string(col_index_from_end)

# エクセルファイルで上から数えて無視する行数
print("エクセルファイルで上から数えて無視する行数を入力してください: ", end="")
excel_row_offset = int(input())

# エクセル => スプレッドシート or スプレッドシート => エクセル
print('エクセル => スプレッドシートなら"e"，スプレッドシート => エクセルなら"s"と入力してください:', end="")
direction = input()


# 取得するエクセルファイルのworkbook, worksheetにアクセス
# 計算結果がほしいのでdata_only=True
excel_wb = openpyxl.load_workbook(excel_file_path, data_only=True)
excel_ws = excel_wb[excel_wb.sheetnames[0]]

# エクセルファイルから学生証番号を取得
excel_student_id_list = excel_ws[excel_student_id_col_index]

# エクセルファイルから参照するcellのrangeを取得
# [1行目, 2行目，...]といった配列
excel_range = excel_ws[(col_index_from_start + "1"):(col_index_from_end + str(excel_ws.max_row))]


#Google Spread Sheetsにアクセス
def connect_gspread(jsonf,key):
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name(jsonf, scope)
    gc = gspread.authorize(credentials)
    SPREADSHEET_KEY = key
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    return workbook

# workbook, worksheetを取得
gspread_wb = connect_gspread(jsonf,spread_sheet_key)
gspread_ws = gspread_wb.get_worksheet(sheet_num)

# スプレッドシートの行数
gspread_row_count = gspread_ws.row_count

# 学生証番号を取得
gspread_student_id_list = gspread_ws.col_values(gspread_student_id_col_num)

# スプレッドシートから参照するcellのリストを取得
# 1行目 => 2行目 => ... という並び順
gspread_score_cell_list = gspread_ws.range(1, col_num_to_start, gspread_row_count, col_num_to_end)

def convert_1d_to_2d(l, cols):
    return [l[i:i + cols] for i in range(0, len(l), cols)]

# 2次元リストに変換
gspread_score_cell_list_2d = convert_1d_to_2d(gspread_score_cell_list, col_num_to_end - col_num_to_start + 1)


# 学生証番号をハイフンで区切る
def format_student_id(raw_student_id):
    raw_student_id = str(raw_student_id)
    return raw_student_id[0:2] + "-" + raw_student_id[2:]

# データの格納
for gspread_row_id in range(gspread_row_offset, len(gspread_student_id_list)):
  gspread_student_id = gspread_student_id_list[gspread_row_id]
  for excel_row_id in range(excel_row_offset, len(excel_student_id_list)):
    # excel_student_id = excel_student_id_list[excel_row_id].value
    excel_student_id = format_student_id(str(excel_student_id_list[excel_row_id].value))
    if gspread_student_id == excel_student_id:
      gspread_cell_row = gspread_score_cell_list_2d[gspread_row_id]
      excel_cell_row = excel_range[excel_row_id]
      if direction == "e":
        for col in range(col_num_to_end - col_num_to_start + 1):
          gspread_cell_row[col].value = excel_cell_row[col].value
      else:
        for col in range(col_num_to_end - col_num_to_start + 1):
          excel_cell_row[col].value = gspread_cell_row[col].value


if direction == "e":
  # 平坦化
  gspread_score_cell_list = sum(gspread_score_cell_list_2d, [])
  gspread_ws.update_cells(gspread_score_cell_list)
else:
  excel_wb.save(excel_file_path)




