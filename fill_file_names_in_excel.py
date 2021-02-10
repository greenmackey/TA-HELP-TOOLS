# フィードバックフォルダに入っているファイル名をエクセルファイルの該当箇所に入力
# フィードバックファイルはPDFを検討しています．

import os
import glob
import openpyxl

# フィードバックファイルを格納しているディレクトリのパス
print("フィードバックファイルを格納しているディレクトリの絶対パス:")
feedback_directory_path = input()

# 書き込む先のエクセルファイルの絶対パス
# 検索すれば絶対パスを簡単に取得する方法が見つかります
print("書き込み先のエクセルファイルの絶対パスを入力してください（e.g., /Users/hoge/Desktop/coremicro/hogehoge.xlsx) :")
excel_file_path = input()

# 全員に共通のコメント
print("全員に共通のコメントを入力してください．なければEnter: ", end="")
base_comment = input() + " "

# 中央値を全員にフィードバックするか
print('中央値を全員にフィードバックしますか？ YES=>y, NO=>n: ', end="")
median_feedback = input()

#ファイル名のリスト取得
file_names = [os.path.basename(file) for file in glob.glob(feedback_directory_path + "/*.pdf") ]

#学生証番号に重複がないか確認
count = 0
seen = set()
for file_idx in range(len(file_names) - 1):
    # 学生証番号を取得（ファイル名の最初の9文字）
    file_name = file_names[file_idx][0:9]
    for file_idx2 in range(file_idx + 1, len(file_names)):
        if file_names[file_idx2][0:9] == file_name:
            if count == 0:
                print("以下の学生証番号でファイル名の重複があります")
                count += 1
            if file_name not in seen:
                print(file_name)
                seen.add(file_name)
if count > 0:
    exit()

#ファイル名に変な記号を含んでいないかチェック
black_list = '¥ / : * ? " < > | % # ` { } ^ [ ]'.split()
for file_name in file_names:
    for black_letter in black_list:
        if black_letter in file_name:
            if count == 0:
                print("以下の学生証番号で不適切な文字列が含まれています")
                count += 1
            print(file_name)
            break
if count > 0:
    exit()


#Workbook, Worksheetを取得
wb = openpyxl.load_workbook(excel_file_path)
sheet = wb[wb.sheetnames[0]]


## medianの計算
## 厳密にはscoreが0より大きい人のmedian
if median_feedback == "y":
    import numpy as np
    score_list = np.array([int(x.value) for x in sheet["F"][1:]\
                            if (x.value != None) and int(x.value) > 0])
    median = np.median(score_list)
    base_comment += "Median: {0} ".format(median)

for row in range(2, sheet.max_row + 1):
    student_id = sheet.cell(row=row, column=1).value
    for file_name in file_names:
        if student_id == file_name[0:9]:
            # ファイル名の入力
            sheet.cell(row=row, column=8).value = file_name
            comment = base_comment
            ## 提出ファイルに問題があった人へののコメント（optional）
            ## pdfフォーマットで提出していなかった人のファイルには"notpdf"という文字列をファイル名に加えておく
            ## ファイルが一つだけでなかった人のファイルには"multiple_files"という文字列をファイル名に加えておく
            if "notpdf" in file_name:
                comment +=  "Submit your file in pdf! "
            if "multiple" in file_name:
                comment += "Submit only one file in pdf! "
            if "invalid_string" in file_name:
                comment += 'Don\'t use ¥ / : * ? " < > | % # ` { } ^ [ ]'
            sheet.cell(row=row, column=7).value = comment
            break

wb.save(excel_file_path)