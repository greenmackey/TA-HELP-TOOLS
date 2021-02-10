# スプレッドシートからスコアを取得し，エクセルファイルの該当箇所に入力
# 必ず，このディレクトリに移動してからコマンドを打つこと

# scrapboxのメモの参考記事にある通り，jsonを当該ディレクトリにダウンロードし，spread_sheet_keyを変更する．

import gspread
from oauth2client.service_account import ServiceAccountCredentials

import openpyxl

# スプレッドシートの鍵を取得
# （例）"10tTyg...5p8k"などの文字列
# jsonとかよくわからんかったら，spread_sheet_keyに値をそのままいれてもいいです．
import json
json_open = open('gspread_info.json', 'r')
json_load = json.load(json_open)
spread_sheet_key = json_load["GSPREAD_SHEET_KEY"]

# jsonfile名を取得
# （例）"coremicro-ii-grade-88c...046.json"などの文字列
# jsonとかよくわからなかったら，jsonfに値をそのままいれてもいいです．
jsonf = json_load["GSPREAD_JSON_FILE_NAME"]


# スプレッドシート内でスコアが入力されている列番号（アルファベット）
# （例）A
print("スプレッドシート内でスコアが入力されている列番号を入力してください（e.g., B）: ", end="")
# アルファベットを数字に直している
col_num = openpyxl.utils.column_index_from_string(input())

# 取得するシート番号（1スタート）
# （例）1
print("取得するシートが何枚目ですか？半角で入力してください: ", end="")
sheet_num = int(input()) - 1

# 書き込む先のエクセルファイルの絶対パス
# 検索すれば絶対パスを簡単に取得する方法が見つかります
print("書き込み先のエクセルファイルの絶対パスを入力してください（e.g., /Users/hoge/Desktop/coremicro/hogehoge.xlsx) :")
excel_file_path = input()


#Google Spread Sheetsにアクセス
def connect_gspread(jsonf,key):
    scope = ['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name(jsonf, scope)
    gc = gspread.authorize(credentials)
    SPREADSHEET_KEY = key
    workbook = gc.open_by_key(SPREADSHEET_KEY)
    return workbook

# workbookを取得
wb = connect_gspread(jsonf,spread_sheet_key)

# 指定されたworksheetを取得
ws = wb.get_worksheet(sheet_num)

# 学生証番号の値リストを取得
student_id_list = ws.col_values(1)

# scoreのリストを取得
score_list = ws.col_values(col_num)

#更新するエクセルファイルのworkbook, worksheetにアクセス
excel_wb = openpyxl.load_workbook(excel_file_path)
excel_ws = excel_wb[excel_wb.sheetnames[0]]

#スコアを入力し保存
for score_idx in range(1, len(student_id_list)):
    for row_idx in range(2, excel_ws.max_row + 1):
        if excel_ws.cell(row=row_idx, column=1).value == student_id_list[score_idx]:
            excel_ws.cell(row=row_idx, column=6).value = score_list[score_idx]
excel_wb.save(excel_file_path)
